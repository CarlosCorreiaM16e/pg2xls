#!/usr/bin/python
# -*- coding: utf-8 -*-

import getopt
import psycopg2
import sys
import os
import traceback

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HELP = """
USAGE: export table to xls

EXAMPLE: python export_table_to_xls.py -d db_name \
[-h localhost] [-p 5432] [-U username] [-W password] \
[-t table] [-q "query"] [-f filename] [-T "sheet title" ]
"""

#------------------------------------------------------------------
def get_column_names_from_query( cursor ):
    column_names = [desc[0] for desc in cursor.description]
    return column_names


#------------------------------------------------------------------
def export_to_xls( conn, table_name, query, title, filename ):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = title
    cursor = conn.cursor()
    if table_name:
        cursor.execute( 'select * from %s order by 1' % table_name )
    else:
        cursor.execute( query )
    col_names = [ c.capitalize().replace( '_', ' ' )
                  for c in get_column_names_from_query( cursor ) ]
    print( col_names )
    ws1.append( col_names )
    for col in range( len( col_names ) ):
        # print( 'col: %d - %s' % (col, get_column_letter( col + 1 ) ) )
        ws1[ '%s1' % get_column_letter( col + 1 ) ].font = Font( bold=True )

    for row in cursor:
        ws1.append( row )
    if not filename.endswith( '.xlsx' ):
        filename = '%s.xlsx' % filename
    wb.save( filename )
    print( 'Stylesheet saved as: %s' % filename )


#------------------------------------------------------------------
def get_password( username ):
    try:
        f = open( '%s/.pgpass' % os.environ[ 'HOME' ], "r" )
        lines = f.readlines()
        f.close()
        for l in lines:
            flds = l.split( ':' )
            if username == flds[ 3 ]:
                password = flds[ 4 ].strip()
                break

    except:
        t, v, tb = sys.exc_info()
        traceback.print_exception( t, v, tb )
        password = raw_input( 'Password: ' )
    return password


#------------------------------------------------------------------
# main
if __name__ == "__main__":
    if len( sys.argv ) < 4 or sys.argv[ 1 ] == '--help':
        print HELP
    else:
        options, args = getopt.getopt( sys.argv[ 1 : ],
                                       "d:h:p:U:W:t:q:f:T:" )
        username = os.environ[ 'USER' ]
        query = None
        table_name = None
        db_name = username
        host = 'localhost'
        port = 5432
        user = username
        title = ''
        filename = None
        password = get_password( username )
        for op, arg in options:
            if op == '-d':
                db_name = arg
            elif op == '-h':
                host = arg
            elif op == '-p':
                port = arg
            elif op == '-U':
                user = arg
            elif op == '-W':
                password = arg
            elif op == '-t':
                table_name = arg
                if not title:
                    title = table_name.capitalize().replace( '_', ' ' )
            elif op == '-q':
                query = arg
            elif op == '-f':
                filename = arg
            elif op == '-T':
                title = arg

        if table_name is None and query is None:
            print( 'Must specify either table_name or query' )
        else:
            # Make the database connection (change driver if required)
            conn = psycopg2.connect( database=db_name,
                                     host=host,
                                     port=port,
                                     user=user,
                                     password=password )
            export_to_xls( conn, table_name, query, title, filename )

